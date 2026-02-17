"""
Manager para Indicação em Massa - Preenche planilha Excel com múltiplos indicados
"""

import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from typing import List, Dict
from datetime import datetime


class IndicacaoMassaManager:
    """Gerencia a criação de planilha de indicação em massa"""
    
    def __init__(self):
        self.template_path = "ficplanilha.xlsx"
        self.mapeamento_funcoes = {
            'S': 'SUPERVISOR',
            'I': 'INSTRUTOR',
            'O': 'OPERADOR',
            'F': 'FMC',
            'CHEQ': 'CHEFE DE EQUIPE',
            'E': 'ESTAGIÁRIO',
            'S/H': 'SEM HABILITAÇÃO',
            '--': 'CHEFE DO COP',
            'SUPERVISOR': 'SUPERVISOR',
            'INSTRUTOR': 'INSTRUTOR',
            'OPERADOR': 'OPERADOR',
            'FMC': 'FMC',
            'CHEFE DE EQUIPE': 'CHEFE DE EQUIPE',
            'ESTAGIÁRIO': 'ESTAGIÁRIO',
            'SEM HABILITAÇÃO': 'SEM HABILITAÇÃO',
        }
    
    def preencher_planilha(self, dados_curso: Dict, indicados: List[Dict], dados_chefes: Dict = None) -> BytesIO:
        """
        Preenche a planilha de indicação em massa
        """
        # Carregar template
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        # Preencher dados do curso (cabeçalho)
        self._preencher_cabecalho(ws, dados_curso)
        
        # Preencher indicados nas linhas 14-35 (22 indicados no template)
        # Se tiver mais de 22, inserir linhas extras
        num_indicados = len(indicados)
        if num_indicados > 22:
            # Inserir linhas extras antes da linha 36
            linhas_extra = num_indicados - 22
            ws.insert_rows(36, linhas_extra)
            # Copiar formato da linha 35 para as novas linhas
            for i in range(36, 36 + linhas_extra):
                self._copiar_formato_linha(ws, 35, i)
        
        # Preencher os dados dos indicados
        self._preencher_indicados(ws, indicados)
        
        # Calcular onde começam as assinaturas
        # Linha 36 + extras (se houver)
        linha_assinaturas = 36 + max(0, num_indicados - 22)
        
        # Preencher seção de assinaturas
        self._preencher_assinaturas(ws, dados_chefes, linha_assinaturas)
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _copiar_formato_linha(self, ws, linha_origem, linha_destino):
        """Copia o formato de uma linha para outra"""
        for col in range(1, 20):
            cel_origem = ws.cell(row=linha_origem, column=col)
            cel_destino = ws.cell(row=linha_destino, column=col)
            if cel_origem.has_style:
                cel_destino.font = cel_origem.font.copy()
                cel_destino.border = cel_origem.border.copy()
                cel_destino.fill = cel_origem.fill.copy()
                cel_destino.number_format = cel_origem.number_format
                cel_destino.protection = cel_origem.protection.copy()
                cel_destino.alignment = cel_origem.alignment.copy()
    
    def _preencher_cabecalho(self, ws, dados_curso: Dict):
        """Preenche o cabeçalho da planilha com dados do curso"""
        # Linha 6: Dados principais
        ws['A6'] = dados_curso.get('codigo', '').upper()
        ws['D6'] = dados_curso.get('nome', '').upper()
        ws['O6'] = 'CRCEA-SE'
        ws['P6'] = 'SIM'  # Previsto no PDP
        ws['Q6'] = dados_curso.get('comando', 'DECEA').upper()
        
        # Linha 9: Turma, local, modalidade, datas
        ws['A9'] = dados_curso.get('turma', '').upper()
        ws['E9'] = dados_curso.get('local', '').upper()
        ws['K9'] = dados_curso.get('modalidade', 'PRESENCIAL').upper()
        ws['N9'] = dados_curso.get('data_inicio', '')
        ws['Q9'] = dados_curso.get('data_termino', '')
    
    def _preencher_indicados(self, ws, indicados: List[Dict]):
        """Preenche os dados dos indicados na planilha (linhas 14-35 ou mais)"""
        # Linha inicial para indicados
        linha_inicial = 14
        
        for idx, pessoa in enumerate(indicados):
            linha = linha_inicial + idx
            
            # Coluna A: Prioridade (1.0, 2.0, etc)
            ws[f'A{linha}'] = float(idx + 1)
            
            # Coluna B: Posto + Nome completo
            nome_completo = self._montar_nome_completo(pessoa)
            ws[f'B{linha}'] = nome_completo
            ws[f'B{linha}'].font = Font(bold=True, size=10)
            
            # Coluna H: CPF formatado
            cpf = self._formatar_cpf(pessoa.get('cpf', ''))
            ws[f'H{linha}'] = cpf
            
            # Coluna L: SARAM
            ws[f'L{linha}'] = self._formatar_saram(pessoa.get('saram', ''))
            
            # Coluna M: Tempo de serviço
            tempo = self._calcular_tempo_servico(pessoa.get('data_praca', ''))
            ws[f'M{linha}'] = tempo
            
            # Coluna N: Função antes do curso (nome completo)
            funcao_atual = pessoa.get('funcao_atual', '')
            funcao_antes = self._get_funcao_completa(funcao_atual)
            ws[f'N{linha}'] = funcao_antes.upper()
            
            # Coluna O: Função depois do curso (nome completo)
            funcao_depois = pessoa.get('funcao_apos_curso', funcao_antes)
            if funcao_depois != funcao_antes:
                funcao_depois = self._get_funcao_completa(funcao_depois)
            ws[f'O{linha}'] = funcao_depois.upper()
            
            # Coluna P: Email funcional
            ws[f'P{linha}'] = (pessoa.get('email', '') or '').upper()
            
            # Coluna R: Celular/Telefone
            ws[f'R{linha}'] = (pessoa.get('telefone', '') or '').upper()
    
    def _preencher_assinaturas(self, ws, dados_chefes: Dict, linha_inicial: int):
        """Preenche a seção de assinaturas"""
        data_atual = datetime.now().strftime('%d/%m/%Y')
        
        # Linha 1: Texto de comunicação + PARECER
        ws[f'A{linha_inicial}'] = 'OS INDICADOS FORAM COMUNICADOS A RESPEITO DESTA INDICAÇÃO E CONFIRMARAM NÃO ESTAREM DE FÉRIAS OU NÃO TEREM OUTROS IMPEDIMENTOS (EX. CURSO CONCOMITANTE QUE NÃO OS DE CARREIRA)'
        ws[f'M{linha_inicial}'] = 'PARECER DO CHEFE DA DIVISÃO/SEÇÃO DO CURSO'
        ws[f'Q{linha_inicial}'] = '( x ) FAVORÁVEL      (    ) DESFAVORÁVEL'
        
        # Linha 2: DATA
        linha_data = linha_inicial + 1
        ws[f'A{linha_data}'] = f'DATA:  {data_atual}'
        ws[f'M{linha_data}'] = f'DATA: {data_atual}'
        
        # Linha 3: Linhas de assinatura
        linha_linha = linha_inicial + 2
        ws[f'A{linha_linha}'] = '__________________________________________________'
        ws[f'N{linha_linha}'] = '__________________________________________________'
        
        # Linha 4: Nomes dos chefes
        linha_nome = linha_inicial + 3
        if dados_chefes and dados_chefes.get('chefe_orgao'):
            nome_chefe = dados_chefes['chefe_orgao'].get('nome', '')
            posto_chefe = dados_chefes['chefe_orgao'].get('posto', '')
            ws[f'A{linha_nome}'] = f"{nome_chefe} {posto_chefe}".strip().upper()
        
        if dados_chefes and dados_chefes.get('chefe_divisao'):
            nome_resp = dados_chefes['chefe_divisao'].get('nome', '')
            posto_resp = dados_chefes['chefe_divisao'].get('posto', '')
            ws[f'N{linha_nome}'] = f"{nome_resp} {posto_resp}".strip().upper()
        
        # Linha 5: Setores
        linha_setor = linha_inicial + 4
        if dados_chefes and dados_chefes.get('chefe_orgao'):
            setor_chefe = dados_chefes['chefe_orgao'].get('setor', '')
            if setor_chefe:
                setor_upper = setor_chefe.strip().upper()
                if setor_upper.startswith('A') or setor_upper.startswith('Ã'):
                    prefixo = 'CHEFE DA'
                else:
                    prefixo = 'CHEFE DO'
                ws[f'A{linha_setor}'] = f"{prefixo} {setor_upper}"
        
        if dados_chefes and dados_chefes.get('chefe_divisao'):
            setor_resp = dados_chefes['chefe_divisao'].get('setor', '')
            if setor_resp:
                setor_upper = setor_resp.strip().upper()
                if setor_upper.startswith('A') or setor_upper.startswith('Ã'):
                    prefixo = 'CHEFE DA'
                else:
                    prefixo = 'CHEFE DO'
                ws[f'N{linha_setor}'] = f"{prefixo} {setor_upper}"
    
    def _get_funcao_completa(self, habilitacao: str) -> str:
        """Retorna o nome completo da função baseado na habilitação/sigla"""
        if not habilitacao:
            return ''
        hab_upper = str(habilitacao).strip().upper()
        return self.mapeamento_funcoes.get(hab_upper, hab_upper)
    
    def _montar_nome_completo(self, pessoa: Dict) -> str:
        """Monta o nome completo com posto"""
        partes = []
        
        if pessoa.get('posto_graduacao'):
            partes.append(str(pessoa['posto_graduacao']).strip().upper())
        
        if pessoa.get('especialidade'):
            partes.append(str(pessoa['especialidade']).strip().upper())
        
        if pessoa.get('nome_completo'):
            partes.append(str(pessoa['nome_completo']).strip().upper())
        
        return ' '.join(partes)
    
    def _formatar_cpf(self, cpf: str) -> str:
        """Formata CPF no padrão XXX.XXX.XXX-XX"""
        if not cpf:
            return ''
        
        import re
        numeros = re.sub(r'\D', '', str(cpf))
        
        if len(numeros) == 11:
            return f"{numeros[:3]}.{numeros[3:6]}.{numeros[6:9]}-{numeros[9:]}"
        
        return cpf
    
    def _formatar_saram(self, saram: str) -> str:
        """Formata SARAM (remove caracteres não numéricos)"""
        if not saram:
            return ''
        
        import re
        return re.sub(r'\D', '', str(saram))
    
    def _calcular_tempo_servico(self, data_praca: str) -> str:
        """Calcula tempo de serviço em anos e meses"""
        if not data_praca:
            return ''
        
        try:
            formatos = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']
            data_praca_dt = None
            
            for formato in formatos:
                try:
                    data_praca_dt = datetime.strptime(str(data_praca).strip(), formato)
                    break
                except ValueError:
                    continue
            
            if not data_praca_dt:
                return ''
            
            hoje = datetime.now()
            anos = hoje.year - data_praca_dt.year
            meses = hoje.month - data_praca_dt.month
            
            if meses < 0:
                anos -= 1
                meses += 12
            
            return f'{anos} anos e {meses} meses'
        except:
            return ''


# Singleton
_indicacao_massa_manager = None

def get_indicacao_massa_manager() -> IndicacaoMassaManager:
    """Retorna a instância singleton do IndicacaoMassaManager"""
    global _indicacao_massa_manager
    if _indicacao_massa_manager is None:
        _indicacao_massa_manager = IndicacaoMassaManager()
    return _indicacao_massa_manager
